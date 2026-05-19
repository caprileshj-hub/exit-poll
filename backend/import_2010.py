"""Import 2010 Asamblea Nacional study into historico tables.

The source workbook has study outputs for nominal, list, and indigenous seats.
There is no local official tabulation file, so the official national reference
uses the published aggregate from Wikipedia/CNE summary.
"""

import json
import sqlite3
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
DB = BASE_DIR / "exitpoll.db"
XLSM = BASE_DIR / "data" / "2010" / "aplicacion03.xlsm"
REF = "2010-asamblea"
NOMBRE = "Asamblea Nacional 2010"
FECHA = "2010-09-26"
TOTAL_ESCANOS = 165

NAME_TO_CODE = {
    "AMAZONAS": "22",
    "ANZOATEGUI": "02",
    "ANZOÁTEGUI": "02",
    "APURE": "03",
    "ARAGUA": "04",
    "BARINAS": "05",
    "BOLIVAR": "06",
    "BOLÍVAR": "06",
    "CARABOBO": "07",
    "COJEDES": "08",
    "DELTA AMACURO": "23",
    "DISTRITO CAPITAL": "01",
    "FALCON": "09",
    "FALCÓN": "09",
    "GUARICO": "10",
    "GUÁRICO": "10",
    "LARA": "11",
    "MERIDA": "12",
    "MÉRIDA": "12",
    "MIRANDA": "13",
    "MONAGAS": "14",
    "NUEVA ESPARTA": "15",
    "PORTUGUESA": "16",
    "SUCRE": "17",
    "TACHIRA": "18",
    "TÁCHIRA": "18",
    "TRUJILLO": "19",
    "VARGAS": "24",
    "YARACUY": "20",
    "ZULIA": "21",
}


def _norm_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value.strip().upper())
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _state_code(value: str | None) -> str | None:
    if not value:
        return None
    return NAME_TO_CODE.get(value) or NAME_TO_CODE.get(_norm_name(value))


def _bando(value) -> str | None:
    if not isinstance(value, str):
        return None
    text = value.strip().lower()
    if "gobierno" in text:
        return "gov"
    if "opos" in text:
        return "opos"
    if "otro" in text:
        return "otros"
    return None


def _pct(part: int, total: int) -> float:
    return round(part * 100.0 / total, 2) if total else 0.0


def _count_unique_centers(wb) -> int:
    centers = set()
    for row in wb["ENTRADA"].iter_rows(min_row=2, values_only=True):
        if row[1]:
            centers.add(row[1])
    return len(centers)


def _seat_counts(wb) -> tuple[Counter, dict[str, Counter]]:
    total = Counter()
    by_state: dict[str, Counter] = defaultdict(Counter)

    def add(sheet: str, bando_cols: list[int]) -> None:
        current_state = None
        for row in wb[sheet].iter_rows(min_row=5, values_only=True):
            if row[1]:
                current_state = str(row[1]).strip().upper()
            for idx in bando_cols:
                value = row[idx] if idx < len(row) else None
                bando = _bando(value)
                if not bando:
                    continue
                total[bando] += 1
                if current_state:
                    by_state[current_state][bando] += 1

    add("R Nominal", [6])
    add("R Lista", [4, 6, 8])
    add("R Indigena", [6])
    return total, by_state


def _study_list_vote_pct(wb) -> dict[str, dict]:
    counts = Counter()
    by_state: dict[str, Counter] = defaultdict(Counter)
    current_state = None
    for row in wb["LISTA"].iter_rows(values_only=True):
        label = row[2]
        total = row[4]
        if isinstance(label, str) and label.startswith("LISTA "):
            current_state = _state_code(label.replace("LISTA ", "").strip().upper())
            continue
        if not current_state or not isinstance(label, str) or not isinstance(total, (int, float)):
            continue
        if label.strip().lower() == "total":
            current_state = None
            continue
        bando = _bando(label) or "otros"
        counts[bando] += float(total)
        if current_state:
            by_state[current_state][bando] += float(total)
    total_votes = sum(counts.values())
    national = {key: round(value * 100.0 / total_votes, 2) for key, value in counts.items()}
    states = {}
    for state_code, state_counts in by_state.items():
        state_total = sum(state_counts.values())
        states[state_code] = {
            key: round(value * 100.0 / state_total, 2)
            for key, value in state_counts.items()
        }
    return {"national": national, "states": states}


OFFICIAL_LISTA_ESTADOS = {
    "AMAZONAS": {"gov": 42.02, "opos": 14.17, "otros": 41.61},
    "ANZOATEGUI": {"gov": 44.93, "opos": 52.24, "otros": 0.86},
    "APURE": {"gov": 60.14, "opos": 36.89, "otros": 1.64},
    "ARAGUA": {"gov": 50.27, "opos": 46.52, "otros": 0.84},
    "BARINAS": {"gov": 56.36, "opos": 42.17, "otros": 0.89},
    "BOLIVAR": {"gov": 50.31, "opos": 47.69, "otros": 0.93},
    "CARABOBO": {"gov": 43.04, "opos": 53.66, "otros": 0.74},
    "COJEDES": {"gov": 63.89, "opos": 32.57, "otros": 0.85},
    "DELTA AMACURO": {"gov": 68.70, "opos": 25.15, "otros": 0.89},
    "DISTRITO CAPITAL": {"gov": 47.73, "opos": 47.80, "otros": 1.11},
    "FALCON": {"gov": 52.21, "opos": 46.25, "otros": 0.98},
    "GUARICO": {"gov": 58.27, "opos": 29.21, "otros": 11.49},
    "LARA": {"gov": 40.77, "opos": 30.11, "otros": 28.43},
    "MERIDA": {"gov": 48.70, "opos": 50.04, "otros": 0.77},
    "MIRANDA": {"gov": 41.44, "opos": 57.12, "otros": 0.58},
    "MONAGAS": {"gov": 58.68, "opos": 35.39, "otros": 0.59},
    "NUEVA ESPARTA": {"gov": 40.87, "opos": 57.92, "otros": 0.69},
    "PORTUGUESA": {"gov": 63.05, "opos": 32.22, "otros": 2.18},
    "SUCRE": {"gov": 51.26, "opos": 47.51, "otros": 0.76},
    "TACHIRA": {"gov": 42.09, "opos": 56.45, "otros": 0.32},
    "TRUJILLO": {"gov": 62.69, "opos": 35.27, "otros": 1.00},
    "VARGAS": {"gov": 54.82, "opos": 43.31, "otros": 0.84},
    "YARACUY": {"gov": 54.56, "opos": 40.39, "otros": 4.60},
    "ZULIA": {"gov": 44.42, "opos": 54.82, "otros": 0.46},
}


def seed_2010(conn: sqlite3.Connection) -> dict[str, int]:
    wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True, keep_vba=False)
    conn.row_factory = sqlite3.Row
    estados = {r["codigo_cne"]: r["nombre"] for r in conn.execute("SELECT codigo_cne, nombre FROM estados")}

    study_seats, state_seats = _seat_counts(wb)
    unique_centers = _count_unique_centers(wb)
    list_vote_pct = _study_list_vote_pct(wb)

    official_seats = {"gov": 98, "opos": 65, "otros": 2}
    official_votes = {"gov": 5423324, "opos": 5320364, "otros": 353979}
    official_vote_pct = {"gov": 48.13, "opos": 47.22, "otros": 3.14}
    voters = 11097667

    notes = {
        "tipo": "asamblea",
        "metrica": "escanos",
        "sin_tendencia": True,
        "estudio_escanos": dict(study_seats),
        "estudio_total_escanos": sum(study_seats.values()),
        "estudio_voto_lista_pct": list_vote_pct["national"],
        "oficial_escanos": official_seats,
        "oficial_total_escanos": TOTAL_ESCANOS,
        "oficial_votos": official_votes,
        "oficial_voto_pct": official_vote_pct,
        "oficial_fuente": "Wikipedia/CNE agregado nacional",
        "nota": "Eleccion parlamentaria con votos nominales, lista e indigenas; comparacion estadal basada en voto lista publicado por Wikipedia/CNE.",
    }

    conn.execute("""
        INSERT INTO historico_estudios
            (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
             pct_gov, pct_opos, pct_otros, num_centros, fuente, notas)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
            nombre=excluded.nombre,
            nombre_eleccion=excluded.nombre_eleccion,
            fecha_eleccion=excluded.fecha_eleccion,
            pct_gov=excluded.pct_gov,
            pct_opos=excluded.pct_opos,
            pct_otros=excluded.pct_otros,
            num_centros=excluded.num_centros,
            fuente=excluded.fuente,
            notas=excluded.notas,
            updated_at=datetime('now')
    """, (
        REF, "NACIONAL", "Nacional", NOMBRE, FECHA,
        _pct(study_seats["gov"], TOTAL_ESCANOS),
        _pct(study_seats["opos"], TOTAL_ESCANOS),
        _pct(study_seats["otros"], TOTAL_ESCANOS),
        unique_centers,
        "aplicacion03.xlsm",
        json.dumps(notes, ensure_ascii=False),
    ))

    conn.execute("""
        INSERT INTO historico_oficial
            (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
             pct_gov, pct_opos, pct_otros, total_votos, fuente)
        VALUES (?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
            nombre=excluded.nombre,
            nombre_eleccion=excluded.nombre_eleccion,
            fecha_eleccion=excluded.fecha_eleccion,
            pct_gov=excluded.pct_gov,
            pct_opos=excluded.pct_opos,
            pct_otros=excluded.pct_otros,
            total_votos=excluded.total_votos,
            fuente=excluded.fuente,
            updated_at=datetime('now')
    """, (
        REF, "NACIONAL", "Nacional", NOMBRE, FECHA,
        _pct(official_seats["gov"], TOTAL_ESCANOS),
        _pct(official_seats["opos"], TOTAL_ESCANOS),
        _pct(official_seats["otros"], TOTAL_ESCANOS),
        voters,
        "Wikipedia/CNE agregado nacional",
    ))

    state_rows = 0
    for state_name, counts in state_seats.items():
        code = _state_code(state_name)
        if not code or code not in estados:
            continue
        total_state = sum(counts.values())
        state_vote_pct = list_vote_pct["states"].get(code, {})
        conn.execute("""
            INSERT INTO historico_estudios
                (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
                 pct_gov, pct_opos, pct_otros, num_centros, fuente, notas)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
                nombre=excluded.nombre,
                nombre_eleccion=excluded.nombre_eleccion,
                fecha_eleccion=excluded.fecha_eleccion,
                pct_gov=excluded.pct_gov,
                pct_opos=excluded.pct_opos,
                pct_otros=excluded.pct_otros,
                num_centros=excluded.num_centros,
                fuente=excluded.fuente,
                notas=excluded.notas,
                updated_at=datetime('now')
        """, (
            REF, code, estados[code], NOMBRE, FECHA,
            _pct(counts["gov"], total_state),
            _pct(counts["opos"], total_state),
            _pct(counts["otros"], total_state),
            total_state,
            "aplicacion03.xlsm",
            json.dumps({
                "tipo": "asamblea",
                "metrica": "escanos",
                "estudio_escanos": dict(counts),
                "estudio_total_escanos_estado": total_state,
                "estudio_voto_lista_pct": state_vote_pct,
            }, ensure_ascii=False),
        ))
        state_rows += 1

    official_state_rows = 0
    for state_name, pct in OFFICIAL_LISTA_ESTADOS.items():
        code = _state_code(state_name)
        if not code or code not in estados:
            continue
        conn.execute("""
            INSERT INTO historico_oficial
                (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
                 pct_gov, pct_opos, pct_otros, total_votos, fuente)
            VALUES (?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
                nombre=excluded.nombre,
                nombre_eleccion=excluded.nombre_eleccion,
                fecha_eleccion=excluded.fecha_eleccion,
                pct_gov=excluded.pct_gov,
                pct_opos=excluded.pct_opos,
                pct_otros=excluded.pct_otros,
                total_votos=excluded.total_votos,
                fuente=excluded.fuente,
                updated_at=datetime('now')
        """, (
            REF, code, estados[code], NOMBRE, FECHA,
            pct["gov"], pct["opos"], pct["otros"],
            0,
            "Wikipedia/CNE voto lista por entidad federal",
        ))
        official_state_rows += 1

    conn.commit()
    return {
        "estudio_filas": state_rows + 1,
        "oficial_filas": official_state_rows + 1,
        "estudio_escanos_gov": study_seats["gov"],
        "estudio_escanos_opos": study_seats["opos"],
        "estudio_escanos_otros": study_seats["otros"],
    }


def main() -> None:
    conn = sqlite3.connect(DB)
    try:
        stats = seed_2010(conn)
    finally:
        conn.close()
    print("2010-asamblea:", stats)


if __name__ == "__main__":
    main()
