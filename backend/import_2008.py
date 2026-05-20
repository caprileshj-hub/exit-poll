"""
Import: Gobernadores 2008 — 25 exit polls regionales.
Lee 7 archivos Excel regionales de data/2008/ y carga:
  - historico_estudios: resultados del exit poll por estado (ambito=slug)
  - historico_oficial: resultados CNE (Wikipedia + PDF municipal)
  - historico_estudios_turnos: serie acumulativa por estado y turno

Uso: python import_2008.py [--reset]
"""
from __future__ import annotations
import os, sys, sqlite3, json, math, re, unicodedata
import openpyxl
from collections import defaultdict

DB   = "exitpoll.db"
BASE = "data/2008"
REF  = "2008-gobernadores"

# ── Slugs canónicos ────────────────────────────────────────────────────────────
SLUG_MAP = {
    "Trujillo": "trujillo", "Táchira": "tachira", "Tachira": "tachira",
    "Mérida": "merida", "Merida": "merida", "Barinas": "barinas",
    "A Mayor": "alcaldia-mayor", "Miranda": "miranda", "Vargas": "vargas",
    "Libertador": "libertador", "Aragua": "aragua", "Cojedes": "cojedes",
    "Carabobo": "carabobo", "Falcón": "falcon", "Falcon": "falcon",
    "Lara": "lara", "Portuguesa": "portuguesa", "Yaracuy": "yaracuy",
    "Apure": "apure", "Bolivar": "bolivar", "Delta Amacuro": "delta-amacuro",
    "Guarico": "guarico", "Guárico": "guarico", "Nueva Esparta": "nueva-esparta",
    "Monagas": "monagas", "Sucre": "sucre", "Anzoategui": "anzoategui",
    "Anzoátegui": "anzoategui", "Maracaibo": "maracaibo", "Zulia": "zulia",
}

NOMBRE_DISPLAY = {
    "trujillo": "Trujillo", "tachira": "Táchira", "merida": "Mérida",
    "barinas": "Barinas", "alcaldia-mayor": "Alcaldía Mayor de Caracas",
    "miranda": "Miranda", "vargas": "Vargas", "libertador": "Municipio Libertador",
    "aragua": "Aragua", "cojedes": "Cojedes", "carabobo": "Carabobo",
    "falcon": "Falcón", "lara": "Lara", "portuguesa": "Portuguesa",
    "yaracuy": "Yaracuy", "apure": "Apure", "bolivar": "Bolívar",
    "delta-amacuro": "Delta Amacuro", "guarico": "Guárico",
    "nueva-esparta": "Nueva Esparta", "monagas": "Monagas", "sucre": "Sucre",
    "anzoategui": "Anzoátegui", "maracaibo": "Maracaibo", "zulia": "Zulia",
}

REGION_MAP = {
    "trujillo": "Andina", "tachira": "Andina", "merida": "Andina", "barinas": "Andina",
    "alcaldia-mayor": "Capital", "miranda": "Capital", "vargas": "Capital", "libertador": "Capital",
    "aragua": "Central", "cojedes": "Central", "carabobo": "Central",
    "falcon": "Centro Occidental", "lara": "Centro Occidental",
    "portuguesa": "Centro Occidental", "yaracuy": "Centro Occidental",
    "apure": "Guayana y Llanos", "bolivar": "Guayana y Llanos",
    "delta-amacuro": "Guayana y Llanos", "guarico": "Guayana y Llanos",
    "nueva-esparta": "Nororiental e Insular", "monagas": "Nororiental e Insular",
    "sucre": "Nororiental e Insular", "anzoategui": "Nororiental e Insular",
    "maracaibo": "Zuliana", "zulia": "Zuliana",
}

TIPO_MAP = {
    "alcaldia-mayor": "Alcaldía Mayor", "libertador": "Alcaldía Municipal",
    "maracaibo": "Alcaldía Municipal",
}

# ── Resultados oficiales (Wikipedia/CNE + PDF municipal) ──────────────────────
OFICIALES = {
    "trujillo":      {"o_gov": 59.96, "o_opos": 26.30, "o_otros": 13.74,
                      "ganador_gov": True,
                      "cand_gov": "Hugo Cabezas (PSUV)", "cand_opos": "Enrique Catalán (UNT)"},
    "tachira":       {"o_gov": 48.12, "o_opos": 49.46, "o_otros":  1.28,
                      "ganador_gov": False,
                      "cand_gov": "Leonardo Salcedo (PSUV)", "cand_opos": "César Pérez Vivas (COPEI)"},
    "merida":        {"o_gov": 54.97, "o_opos": 44.76, "o_otros":  0.27,
                      "ganador_gov": True,
                      "cand_gov": "Marcos Díaz Orellana (PSUV)", "cand_opos": "Williams Dávila (AD)"},
    "barinas":       {"o_gov": 50.00, "o_opos": 44.30, "o_otros":  5.70,
                      "ganador_gov": True,
                      "cand_gov": "Adán Chávez (PSUV)", "cand_opos": "Julio César Reyes (coalición)"},
    "alcaldia-mayor":{"o_gov": 44.97, "o_opos": 52.40, "o_otros":  1.97,
                      "ganador_gov": False,
                      "cand_gov": "Aristóbulo Istúriz (PSUV)", "cand_opos": "Antonio Ledezma (PJ)"},
    "miranda":       {"o_gov": 46.10, "o_opos": 53.11, "o_otros":  0.00,
                      "ganador_gov": False,
                      "cand_gov": "Diosdado Cabello (PSUV)", "cand_opos": "Henrique Capriles (PJ)"},
    "vargas":        {"o_gov": 61.57, "o_opos": 32.19, "o_otros":  2.81,
                      "ganador_gov": True,
                      "cand_gov": "Jorge García Carneiro (PSUV)", "cand_opos": "Roberto Smith (VdP)"},
    "libertador":    {"o_gov": 53.59, "o_opos": 41.39, "o_otros":  5.02,
                      "ganador_gov": True,
                      "cand_gov": "Jorge Rodríguez (PSUV)", "cand_opos": "Stalin González (coalición)"},
    "aragua":        {"o_gov": 58.77, "o_opos": 39.96, "o_otros":  1.27,
                      "ganador_gov": True,
                      "cand_gov": "Rafael Isea (PSUV)", "cand_opos": "Henry Rosales (PODEMOS)"},
    "cojedes":       {"o_gov": 52.42, "o_opos": 39.60, "o_otros":  7.98,
                      "ganador_gov": True,
                      "cand_gov": "Teodoro Bolívar (PSUV)", "cand_opos": "Alberto Galíndez (AD)"},
    "carabobo":      {"o_gov": 44.48, "o_opos": 47.54, "o_otros":  6.56,
                      "ganador_gov": False,
                      "cand_gov": "Mario Silva (PSUV)", "cand_opos": "Henrique Salas Feo (PV)"},
    "falcon":        {"o_gov": 55.35, "o_opos": 44.40, "o_otros":  0.25,
                      "ganador_gov": True,
                      "cand_gov": "Stella Lugo (PSUV)", "cand_opos": "Gregorio Graterol (AD)"},
    "lara":          {"o_gov": None,  "o_opos": None, "o_otros": None,
                      "ganador_gov": None,
                      "nota_lara": "Henri Falcón ganó con 73.52% (PPT/AP). En el estudio figura como candidato de gobierno (incumbente PPT). Categorización ambigua: PPT era aliado del gobierno en 2008.",
                      "cand_gov": "Henri Falcón (PPT)", "cand_opos": "Pedro Pablo Alcántara (AD)",
                      "o_gov_ppt": 73.52, "o_opos_ad": 14.58},
    "portuguesa":    {"o_gov": 57.89, "o_opos": 26.94, "o_otros": 14.69,
                      "ganador_gov": True,
                      "cand_gov": "Wilmar Castro Soteldo (PSUV)", "cand_opos": "Jóbito Villegas (AD)"},
    "yaracuy":       {"o_gov": 57.83, "o_opos": 28.91, "o_otros":  9.97,
                      "ganador_gov": True,
                      "cand_gov": "Julio León Heredia (PSUV)", "cand_opos": "Filippo Lapi (Convergencia)"},
    "apure":         {"o_gov": 56.58, "o_opos": 26.65, "o_otros":  7.02,
                      "ganador_gov": True,
                      "cand_gov": "Jesús Aguilarte (PSUV)", "cand_opos": "Miriam de Montilla (AD)"},
    "bolivar":       {"o_gov": 47.03, "o_opos": 30.79, "o_otros": 15.00,
                      "ganador_gov": True,
                      "cand_gov": "Francisco Rangel Gómez (PSUV)", "cand_opos": "Andrés Velásquez (LCR)"},
    "delta-amacuro": {"o_gov": 55.61, "o_opos": 25.37, "o_otros": 14.47,
                      "ganador_gov": True,
                      "cand_gov": "Lizeta Hernández (PSUV)", "cand_opos": "Víctor Cedeño (MERI)"},
    "guarico":       {"o_gov": 52.39, "o_opos": 13.40, "o_otros": 33.37,
                      "ganador_gov": True,
                      "cand_gov": "Willian Lara (PSUV)", "cand_opos": "Reynaldo Armas (AD)"},
    "nueva-esparta": {"o_gov": 41.80, "o_opos": 57.53, "o_otros":  0.67,
                      "ganador_gov": False,
                      "cand_gov": "William Fariñas (PSUV)", "cand_opos": "Morel Rodríguez (AD)"},
    "monagas":       {"o_gov": 64.86, "o_opos": 15.02, "o_otros": 12.88,
                      "ganador_gov": True,
                      "cand_gov": "José Gregorio Briceño (PSUV)", "cand_opos": "Domingo Urbina (AD)"},
    "sucre":         {"o_gov": 56.51, "o_opos": 42.21, "o_otros":  1.26,
                      "ganador_gov": True,
                      "cand_gov": "Enrique Maestre (PSUV)", "cand_opos": "Eduardo Morales Gil (AD)"},
    "anzoategui":    {"o_gov": 55.05, "o_opos": 40.52, "o_otros":  3.34,
                      "ganador_gov": True,
                      "cand_gov": "Tarek William Saab (PSUV)", "cand_opos": "Gustavo Marcano (AD)"},
    "maracaibo":     {"o_gov": 39.71, "o_opos": 59.90, "o_otros":  0.39,
                      "ganador_gov": False,
                      "cand_gov": "Henry Ramírez (PSUV)", "cand_opos": "Manuel Rosales (UNT)"},
    "zulia":         {"o_gov": 45.27, "o_opos": 53.33, "o_otros":  1.36,
                      "ganador_gov": False,
                      "cand_gov": "Gian Carlo Di Martino (PSUV)", "cand_opos": "Pablo Pérez (UNT)"},
}

ICC_REF = 0.04  # Design effect reference ICC

# ── Helpers ────────────────────────────────────────────────────────────────────

def _slug(name: str) -> str:
    return SLUG_MAP.get(name, re.sub(r"[^a-z0-9-]", "", unicodedata.normalize("NFD", name.lower()).encode("ascii", "ignore").decode()))

def _deff(n: int, k: int) -> float | None:
    if k <= 0 or n <= 0:
        return None
    m_bar = n / k
    return round(1 + (m_bar - 1) * ICC_REF, 2)

def _moe(n: int, deff: float | None) -> float | None:
    if not n:
        return None
    d = deff if deff else 1.0
    return round(1.96 * math.sqrt(d * 0.25 / n) * 100, 2)

# ── Identificar hojas gráfico vs estado ────────────────────────────────────────

def is_graphic_sheet(name: str, all_names: list[str]) -> bool:
    for n in all_names:
        if n == name:
            continue
        if name == "G" + n or name == "C" + n:
            return True
    return False

# ── Parsear hoja de transcripción ─────────────────────────────────────────────

def parse_state_sheet(ws) -> dict | None:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    candidates: list[str] = []
    for i, row in enumerate(rows):
        if row[1] == "Entidad":
            header_idx = i
            candidates = [v.strip() for v in row[4:] if v and isinstance(v, str) and v.strip()]
            break
    if header_idx is None:
        return None

    centros: set = set()
    # Map: (centro, turno) -> list of vote arrays (may have duplicates)
    readings: dict[tuple, list[int]] = {}
    max_turno = 0

    for row in rows[header_idx + 1:]:
        if row[2] is None or not isinstance(row[2], (int, float)):
            continue
        centro = int(row[2])
        turno = int(row[3]) if isinstance(row[3], (int, float)) else 1
        centros.add(centro)
        max_turno = max(max_turno, turno)
        votos = [int(row[4 + k]) if isinstance(row[4 + k], (int, float)) else 0
                 for k in range(len(candidates))]
        key = (centro, turno)
        # Keep last reading per (centro, turno)
        readings[key] = votos

    if not readings:
        return None

    # Aggregate totals
    total_votos = [0] * len(candidates)
    for v in readings.values():
        for k, val in enumerate(v):
            total_votos[k] += val
    total = sum(total_votos)
    pct = [round(v / total * 100, 2) if total else 0 for v in total_votos]

    return {
        "candidates": candidates,
        "centros": centros,
        "n_centros": len(centros),
        "max_turno": max_turno,
        "total": total,
        "votos": total_votos,
        "pct": pct,
        "readings": readings,  # needed for turno series
    }

# ── Construir serie de turnos acumulativa ────────────────────────────────────

def build_turno_series(parsed: dict) -> list[dict]:
    """
    Para cada turno T, calcula el porcentaje acumulativo usando
    para cada centro su última lectura con turno <= T.
    """
    readings = parsed["readings"]
    max_turno = parsed["max_turno"]
    candidates = parsed["candidates"]
    n_cands = len(candidates)

    # Organizar por centro: lista de (turno, votos) ordenada
    por_centro: dict[int, list[tuple]] = defaultdict(list)
    for (centro, turno), votos in readings.items():
        por_centro[centro].append((turno, votos))
    for c in por_centro:
        por_centro[c].sort()

    series = []
    for t in range(1, max_turno + 1):
        acum = [0] * n_cands
        centros_activos = 0
        for centro, lecturas in por_centro.items():
            # Última lectura con turno <= t
            ultima = None
            for turno, votos in lecturas:
                if turno <= t:
                    ultima = votos
                else:
                    break
            if ultima is not None:
                centros_activos += 1
                for k, v in enumerate(ultima):
                    acum[k] += v
        total_t = sum(acum)
        if total_t == 0:
            continue
        pct_t = [round(v / total_t * 100, 2) for v in acum]
        series.append({
            "turno": t,
            "pct_gov": pct_t[0] if pct_t else 0,
            "pct_opos": pct_t[1] if len(pct_t) > 1 else 0,
            "pct_otros": round(100 - pct_t[0] - (pct_t[1] if len(pct_t) > 1 else 0), 2),
            "num_centros": centros_activos,
        })
    return series

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    reset = "--reset" in sys.argv
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = OFF")

    if reset:
        print("Eliminando datos previos de 2008-gobernadores...")
        conn.execute("DELETE FROM historico_estudios WHERE eleccion_ref=?", (REF,))
        conn.execute("DELETE FROM historico_oficial  WHERE eleccion_ref=?", (REF,))
        conn.execute("DELETE FROM historico_estudios_turnos WHERE eleccion_ref=?", (REF,))
        conn.commit()

    estados: list[dict] = []

    for fn in sorted(os.listdir(BASE)):
        if not fn.endswith(".xlsx"):
            continue
        path = os.path.join(BASE, fn)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        all_sheets = wb.sheetnames
        state_sheets = [s for s in all_sheets
                        if s not in {"PPrincipal"} and not is_graphic_sheet(s, all_sheets)]
        for sheet_name in state_sheets:
            slug = _slug(sheet_name)
            parsed = parse_state_sheet(wb[sheet_name])
            if not parsed:
                print(f"  [SKIP] {sheet_name}: sin datos")
                continue
            estados.append({"slug": slug, "sheet": sheet_name, "parsed": parsed})
        wb.close()

    print(f"Estados parseados: {len(estados)}")

    # ── Fila NACIONAL del estudio (colección) ─────────────────────────────────
    total_centros = sum(e["parsed"]["n_centros"] for e in estados)
    total_resp    = sum(e["parsed"]["total"] for e in estados)
    gov_wins_e    = sum(1 for e in estados if e["parsed"]["pct"] and len(e["parsed"]["pct"]) > 1
                        and e["parsed"]["pct"][0] > e["parsed"]["pct"][1])
    opos_wins_e   = len(estados) - gov_wins_e

    notas_nac = json.dumps({
        "tipo": "coleccion_gobernadores",
        "n_estudios": len(estados),
        "n_centros_total": total_centros,
        "n_respondentes_total": total_resp,
        "icc_referencia": ICC_REF,
        "gov_wins_estudio": gov_wins_e,
        "opos_wins_estudio": opos_wins_e,
        "fuente_oficial": "Wikipedia/CNE + PDF resultados municipales 2008",
    }, ensure_ascii=False)

    conn.execute("""
        INSERT INTO historico_estudios
            (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
             pct_gov, pct_opos, pct_otros, num_centros, fuente, notas)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
            nombre=excluded.nombre, nombre_eleccion=excluded.nombre_eleccion,
            num_centros=excluded.num_centros, notas=excluded.notas,
            updated_at=datetime('now')
    """, (REF, "NACIONAL", "Gobernadores 2008 (colección)",
          "Elecciones Regionales 2008", "2008-11-23",
          0, 0, 0, total_centros, "Cores Excel / importacion", notas_nac))

    # ── Filas por estado ──────────────────────────────────────────────────────
    for e in estados:
        slug   = e["slug"]
        parsed = e["parsed"]
        cands  = parsed["candidates"]
        pct    = parsed["pct"]
        of_    = OFICIALES.get(slug, {})

        # e_gov = col 0, e_opos = col 1, e_otros = resto
        e_gov  = pct[0] if pct else None
        e_opos = pct[1] if len(pct) > 1 else None
        e_otros = round(100 - (e_gov or 0) - (e_opos or 0), 2) if e_gov is not None else None

        deff = _deff(parsed["total"], parsed["n_centros"])
        moe_srs = _moe(parsed["total"], 1.0)
        moe_adj = _moe(parsed["total"], deff)

        notas_estado = json.dumps({
            "tipo": "gobernador",
            "tipo_cargo": TIPO_MAP.get(slug, "Gobernación"),
            "region": REGION_MAP.get(slug, ""),
            "candidato_gov":  cands[0] if cands else None,
            "candidato_opos": cands[1] if len(cands) > 1 else None,
            "candidato_otros": cands[2:] if len(cands) > 2 else [],
            "cand_gov_nombre":  of_.get("cand_gov"),
            "cand_opos_nombre": of_.get("cand_opos"),
            "n_respondentes": parsed["total"],
            "deff_estimado":  deff,
            "moe_srs_pp":     moe_srs,
            "moe_ajustado_pp": moe_adj,
            "lara_nota": of_.get("nota_lara"),
        }, ensure_ascii=False)

        conn.execute("""
            INSERT INTO historico_estudios
                (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
                 pct_gov, pct_opos, pct_otros, num_centros, fuente, notas)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
                nombre=excluded.nombre, pct_gov=excluded.pct_gov,
                pct_opos=excluded.pct_opos, pct_otros=excluded.pct_otros,
                num_centros=excluded.num_centros, notas=excluded.notas,
                updated_at=datetime('now')
        """, (REF, slug, NOMBRE_DISPLAY.get(slug, slug),
              "Elecciones Regionales 2008", "2008-11-23",
              e_gov, e_opos, e_otros, parsed["n_centros"],
              "Cores Excel / importacion", notas_estado))

        # Oficial
        o_gov  = of_.get("o_gov")
        o_opos = of_.get("o_opos")
        o_otros = of_.get("o_otros")
        if o_gov is None and slug == "lara":
            o_gov  = of_.get("o_gov_ppt")
            o_opos = of_.get("o_opos_ad")

        conn.execute("""
            INSERT INTO historico_oficial
                (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
                 pct_gov, pct_opos, pct_otros, fuente)
            VALUES (?,?,?,?,?,?,?,?,?)
            ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
                pct_gov=excluded.pct_gov, pct_opos=excluded.pct_opos,
                pct_otros=excluded.pct_otros, updated_at=datetime('now')
        """, (REF, slug, NOMBRE_DISPLAY.get(slug, slug),
              "Elecciones Regionales 2008", "2008-11-23",
              o_gov or 0, o_opos or 0, o_otros or 0,
              "Wikipedia/CNE + PDF resultados 2008"))

        # Turnos acumulativos
        serie = build_turno_series(parsed)
        for s in serie:
            conn.execute("""
                INSERT INTO historico_estudios_turnos
                    (eleccion_ref, ambito, turno, pct_gov, pct_opos, pct_otros, num_centros)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(eleccion_ref, ambito, turno) DO UPDATE SET
                    pct_gov=excluded.pct_gov, pct_opos=excluded.pct_opos,
                    pct_otros=excluded.pct_otros, num_centros=excluded.num_centros,
                    updated_at=datetime('now')
            """, (REF, slug, s["turno"], s["pct_gov"], s["pct_opos"], s["pct_otros"], s["num_centros"]))

        gov_str  = f"{e_gov}%" if e_gov is not None else "—"
        opos_str = f"{e_opos}%" if e_opos is not None else "—"
        win_e = "GOV " if (e_gov and e_opos and e_gov > e_opos) else "OPOS"
        win_o = "GOV " if of_.get("ganador_gov") else ("OPOS" if of_.get("ganador_gov") is False else "  ? ")
        ok = "OK" if (win_e.strip() == win_o.strip()) else "XX"
        print(f"  {ok} {slug:<18} estudio={gov_str:>6}/{opos_str:<6} {win_e}  "
              f"oficial={of_.get('o_gov','?')}%/{of_.get('o_opos','?')}% {win_o}  "
              f"n={parsed['total']:>6}  k={parsed['n_centros']:>3}  "
              f"T={parsed['max_turno']:>2}  DEFF={deff}")

    conn.commit()
    conn.execute("PRAGMA foreign_keys = ON")
    conn.close()

    # Totales
    print()
    ne = sum(1 for _ in estados) + 1  # +1 NACIONAL
    gov_wins_e = sum(1 for e in estados if e["parsed"]["pct"] and len(e["parsed"]["pct"]) > 1
                     and e["parsed"]["pct"][0] > e["parsed"]["pct"][1])
    gov_wins_o = sum(1 for e in estados if OFICIALES.get(e["slug"], {}).get("ganador_gov") is True)
    correctos  = sum(1 for e in estados
                     if e["parsed"]["pct"] and len(e["parsed"]["pct"]) > 1
                     and OFICIALES.get(e["slug"], {}).get("ganador_gov") is not None
                     and (e["parsed"]["pct"][0] > e["parsed"]["pct"][1]) == OFICIALES[e["slug"]]["ganador_gov"])
    con_oficial = sum(1 for e in estados if OFICIALES.get(e["slug"], {}).get("ganador_gov") is not None)
    print(f"Filas cargadas: {ne} estudio + {len(estados)} oficial + turnos")
    print(f"Estudio: gov gana en {gov_wins_e}/{len(estados)}")
    print(f"Oficial: gov gana en {gov_wins_o}/{con_oficial}")
    print(f"Ganador correcto: {correctos}/{con_oficial}  ({round(correctos/con_oficial*100,1)}%)")

if __name__ == "__main__":
    main()
