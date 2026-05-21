"""Import Presidencial 2012 and 2013 into historico tables.

Los archivos fuente viven bajo subcarpetas por tipo de eleccion, porque 2012 y
2013 tienen mas de un estudio historico cargado en el repo.
"""
import openpyxl
import sqlite3
from collections import defaultdict

DB = 'exitpoll.db'
DATA_2012_PRESIDENCIAL = 'data/2012/presidenciales'
DATA_2013_PRESIDENCIAL = 'data/2013/presidenciales'

# ── Name → DB codigo_cne mapping ─────────────────────────────────────────────
NAME_TO_CODE = {
    'amazonas': '22', 'anzoategui': '02', 'anzoátegui': '02',
    'apure': '03', 'aragua': '04', 'barinas': '05',
    'bolivar': '06', 'bolívar': '06', 'carabobo': '07', 'cojedes': '08',
    'delta amacuro': '23', 'distrito capital': '01',
    'dtto. capital': '01', 'dtto capital': '01',
    'falcon': '09', 'falcón': '09', 'guarico': '10', 'guárico': '10',
    'lara': '11', 'merida': '12', 'mérida': '12', 'miranda': '13',
    'monagas': '14', 'nueva esparta': '15', 'portuguesa': '16',
    'sucre': '17', 'tachira': '18', 'táchira': '18', 'trujillo': '19',
    'vargas': '24', 'yaracuy': '20', 'zulia': '21',
}


def name_to_cod(name: str) -> str | None:
    return NAME_TO_CODE.get(name.strip().lower())


def fv(x) -> float:
    try:
        return float(x)
    except Exception:
        return 0.0


def r2(x: float) -> float:
    return round(x, 2)


def get_conn():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


# ── Upsert helpers ────────────────────────────────────────────────────────────
def upsert_estudio(conn, ref, nombre_eleccion, fecha, ambito, nombre,
                   pct_gov, pct_opos, pct_otros, num_centros=0):
    conn.execute("""
        INSERT INTO historico_estudios
            (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
             pct_gov, pct_opos, pct_otros, num_centros)
        VALUES (?,?,?,?,?,?,?,?,?)
        ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
            nombre=excluded.nombre, nombre_eleccion=excluded.nombre_eleccion,
            fecha_eleccion=excluded.fecha_eleccion,
            pct_gov=excluded.pct_gov, pct_opos=excluded.pct_opos,
            pct_otros=excluded.pct_otros, num_centros=excluded.num_centros,
            updated_at=datetime('now')
    """, (ref, ambito, nombre, nombre_eleccion, fecha,
          pct_gov, pct_opos, pct_otros, num_centros))


def upsert_oficial(conn, ref, nombre_eleccion, fecha, ambito, nombre,
                   pct_gov, pct_opos, pct_otros, total_votos=0):
    conn.execute("""
        INSERT INTO historico_oficial
            (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
             pct_gov, pct_opos, pct_otros, total_votos)
        VALUES (?,?,?,?,?,?,?,?,?)
        ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
            nombre=excluded.nombre, nombre_eleccion=excluded.nombre_eleccion,
            fecha_eleccion=excluded.fecha_eleccion,
            pct_gov=excluded.pct_gov, pct_opos=excluded.pct_opos,
            pct_otros=excluded.pct_otros, total_votos=excluded.total_votos,
            updated_at=datetime('now')
    """, (ref, ambito, nombre, nombre_eleccion, fecha,
          pct_gov, pct_opos, pct_otros, total_votos))


def upsert_turno(conn, ref, turno, pct_gov, pct_opos, pct_otros, num_centros=0):
    conn.execute("""
        INSERT INTO historico_estudios_turnos
            (eleccion_ref, ambito, turno, pct_gov, pct_opos, pct_otros, num_centros)
        VALUES (?,'NACIONAL',?,?,?,?,?)
        ON CONFLICT(eleccion_ref, ambito, turno) DO UPDATE SET
            pct_gov=excluded.pct_gov, pct_opos=excluded.pct_opos,
            pct_otros=excluded.pct_otros, num_centros=excluded.num_centros,
            updated_at=datetime('now')
    """, (ref, turno, pct_gov, pct_opos, pct_otros, num_centros))


# ══════════════════════════════════════════════════════════════════════════════
# 2012 PRESIDENCIAL  (Chávez vs Capriles, 7-Oct-2012)
# ══════════════════════════════════════════════════════════════════════════════
def import_2012(conn):
    REF    = '2012-presidencial'
    NOMBRE = 'Presidencial 2012'
    FECHA  = '2012-10-07'

    wb = openpyxl.load_workbook(f'{DATA_2012_PRESIDENCIAL}/Core (2).xlsx', read_only=True, data_only=True)

    # ── Estado DB nombres ────────────────────────────────────────────────────
    estado_nombre = {r['codigo_cne']: r['nombre']
                     for r in conn.execute("SELECT codigo_cne, nombre FROM estados")}

    # ── Centros por turno (Entrada) ──────────────────────────────────────────
    ws_ent = wb['Entrada']
    by_turno_ctrs = defaultdict(set)
    for row in ws_ent.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        t   = int(fv(row[2]))
        cod = int(fv(row[0]))
        by_turno_ctrs[t].add(cod)
    nat_ctrs = len({c for s in by_turno_ctrs.values() for c in s})

    # ── Turn-by-turn  (VZA sheet, Venezuela rows: Chavez, Capriles, Otro) ───
    ws_vza = wb['VZA']
    vza_rows = list(ws_vza.iter_rows(min_row=2, max_row=4, values_only=True))
    # Row 0 = Chavez (cols 5..24 = turn increments 1..20)
    # Row 1 = Capriles
    # Row 2 = Otro
    ch_inc  = [fv(vza_rows[0][c]) for c in range(5, 25)]
    cap_inc = [fv(vza_rows[1][c]) for c in range(5, 25)]
    ot_inc  = [fv(vza_rows[2][c]) for c in range(5, 25)]

    cum_ch = cum_cap = cum_ot = 0.0
    seen_ctrs: set = set()
    for t in range(1, 21):
        inc_ch  = ch_inc[t - 1]
        inc_cap = cap_inc[t - 1]
        inc_ot  = ot_inc[t - 1]
        if inc_ch == 0 and inc_cap == 0 and inc_ot == 0:
            continue
        cum_ch  += inc_ch
        cum_cap += inc_cap
        cum_ot  += inc_ot
        seen_ctrs |= by_turno_ctrs.get(t, set())
        tot_t = cum_ch + cum_cap + cum_ot
        upsert_turno(conn, REF, t,
                     r2(cum_ch  / tot_t * 100),
                     r2(cum_cap / tot_t * 100),
                     r2(cum_ot  / tot_t * 100),
                     len(seen_ctrs))

    # ── Final study results  (R1: weighted national + per-state) ────────────
    ws_r1 = wb['R1']
    # Header row: ENTIDAD FEDERAL | CAPRILES | CHAVEZ | FACTOR | cap_nat | chav_nat | diff
    nat_cap = nat_chav = None
    for row in ws_r1.iter_rows(min_row=4, values_only=True):
        state = row[3]
        if state is None:
            continue
        state_s = str(state).strip()
        if state_s == 'VENEZUELA':
            # cols [7]=cap_nat, [8]=chav_nat
            nat_cap  = fv(row[7])
            nat_chav = fv(row[8])
        else:
            cap_pct  = fv(row[4])
            chav_pct = fv(row[5])
            otros    = max(0.0, 1.0 - cap_pct - chav_pct)
            cod = name_to_cod(state_s)
            if cod and cod in estado_nombre:
                upsert_estudio(conn, REF, NOMBRE, FECHA, cod, estado_nombre[cod],
                               r2(chav_pct * 100), r2(cap_pct * 100),
                               r2(otros * 100), 0)

    nat_otros = max(0.0, 1.0 - nat_chav - nat_cap)
    upsert_estudio(conn, REF, NOMBRE, FECHA, 'NACIONAL', 'Nacional',
                   r2(nat_chav * 100), r2(nat_cap * 100),
                   r2(nat_otros * 100), nat_ctrs)

    # ── Official results (xlsx per mesa) ─────────────────────────────────────
    # cols: cod_edo(0), chavez(11), capriles(12), chirino(13), sequera(14),
    #       reyes(15), bolivar(16), votos_nulos(17)
    wb_off = openpyxl.load_workbook(
        f'{DATA_2012_PRESIDENCIAL}/resultados oficiales presidenciales 2012.xlsx',
        read_only=True, data_only=True)
    ws_off = wb_off.active
    by_edo = defaultdict(lambda: {'chav': 0, 'cap': 0, 'otros': 0, 'tot': 0})
    nat_off = {'chav': 0, 'cap': 0, 'otros': 0, 'tot': 0}
    for row in ws_off.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        cod_edo = int(row[0])
        if cod_edo == 99:
            continue  # excluir exterior
        chav  = int(row[11] or 0)
        cap   = int(row[12] or 0)
        otros = (int(row[13] or 0) + int(row[14] or 0) +
                 int(row[15] or 0) + int(row[16] or 0) + int(row[17] or 0))
        tot   = chav + cap + otros
        by_edo[cod_edo]['chav']  += chav
        by_edo[cod_edo]['cap']   += cap
        by_edo[cod_edo]['otros'] += otros
        by_edo[cod_edo]['tot']   += tot
        nat_off['chav']  += chav
        nat_off['cap']   += cap
        nat_off['otros'] += otros
        nat_off['tot']   += tot

    d = nat_off['tot']
    upsert_oficial(conn, REF, NOMBRE, FECHA, 'NACIONAL', 'Nacional',
                   r2(nat_off['chav'] / d * 100), r2(nat_off['cap'] / d * 100),
                   r2(nat_off['otros'] / d * 100), d)
    for edo, s in by_edo.items():
        cod = f'{edo:02d}'
        if cod not in estado_nombre or s['tot'] == 0:
            continue
        upsert_oficial(conn, REF, NOMBRE, FECHA, cod, estado_nombre[cod],
                       r2(s['chav'] / s['tot'] * 100),
                       r2(s['cap']  / s['tot'] * 100),
                       r2(s['otros'] / s['tot'] * 100),
                       s['tot'])

    print(f'2012: estudio nacional Chavez={r2(nat_chav*100)}% Capriles={r2(nat_cap*100)}%'
          f' | oficial Chavez={r2(nat_off["chav"]/d*100)}% Capriles={r2(nat_off["cap"]/d*100)}%')


# ══════════════════════════════════════════════════════════════════════════════
# 2013 PRESIDENCIAL  (Maduro vs Capriles, 14-Abr-2013)
# ══════════════════════════════════════════════════════════════════════════════
def import_2013(conn):
    REF    = '2013-presidencial'
    NOMBRE = 'Presidencial 2013'
    FECHA  = '2013-04-14'

    wb = openpyxl.load_workbook(f'{DATA_2013_PRESIDENCIAL}/CoreMA2.xlsx', read_only=True, data_only=True)

    estado_nombre = {r['codigo_cne']: r['nombre']
                     for r in conn.execute("SELECT codigo_cne, nombre FROM estados")}

    # ── Centros por turno (Entrada) ──────────────────────────────────────────
    # cols: transcriptor(1), telefonista(2), codigo(3), turnoC(4), turnoA(5),
    #       Maduro(6), Capriles(7), Otro(8), totalv(9)
    ws_ent = wb['Entrada']
    by_turno_ctrs = defaultdict(set)
    for row in ws_ent.iter_rows(min_row=3, values_only=True):
        if not row[3]:
            continue
        t   = int(fv(row[4]))   # turnoC = turn of data collection
        cod = int(fv(row[3]))
        by_turno_ctrs[t].add(cod)
    nat_ctrs = len({c for s in by_turno_ctrs.values() for c in s})

    # ── Turn-by-turn (Venezuela sheet, incremental vote counts per turn) ─────
    # cols: Candidato(1), Resultado(2), Votos(3), turn1..turn18(4..21)
    ws_ven = wb['Venezuela']
    ven_rows = list(ws_ven.iter_rows(min_row=3, max_row=5, values_only=True))
    # Row 0=Maduro, Row 1=Capriles, Row 2=Otro
    mad_inc = [fv(ven_rows[0][c]) for c in range(4, 22)]   # turns 1-18
    cap_inc = [fv(ven_rows[1][c]) for c in range(4, 22)]
    ot_inc  = [fv(ven_rows[2][c]) for c in range(4, 22)]

    cum_mad = cum_cap = cum_ot = 0.0
    seen_ctrs: set = set()
    for t in range(1, 19):
        inc_mad = mad_inc[t - 1]
        inc_cap = cap_inc[t - 1]
        inc_ot  = ot_inc[t - 1]
        if inc_mad == 0 and inc_cap == 0 and inc_ot == 0:
            continue
        cum_mad += inc_mad
        cum_cap += inc_cap
        cum_ot  += inc_ot
        seen_ctrs |= by_turno_ctrs.get(t, set())
        tot_t = cum_mad + cum_cap + cum_ot
        upsert_turno(conn, REF, t,
                     r2(cum_mad / tot_t * 100),
                     r2(cum_cap / tot_t * 100),
                     r2(cum_ot  / tot_t * 100),
                     len(seen_ctrs))

    # ── Final study results (R1: vote counts by state) ───────────────────────
    # cols: state(1), maduro(2), capriles(3), otros(4), diff(5)
    ws_r1 = wb['R1']
    nat_mad = nat_cap_s = nat_ot_s = 0
    for row in ws_r1.iter_rows(min_row=5, values_only=True):
        state = row[1]
        if not state:
            continue
        state_s = str(state).strip()
        mad  = fv(row[2])
        cap  = fv(row[3])
        ot   = fv(row[4])
        tot  = mad + cap + ot
        if tot == 0:
            continue
        if state_s == 'VENEZUELA':
            nat_mad  = mad
            nat_cap_s = cap
            nat_ot_s  = ot
        else:
            cod = name_to_cod(state_s)
            if cod and cod in estado_nombre:
                upsert_estudio(conn, REF, NOMBRE, FECHA, cod, estado_nombre[cod],
                               r2(mad / tot * 100), r2(cap / tot * 100),
                               r2(ot  / tot * 100), 0)

    nat_tot_s = nat_mad + nat_cap_s + nat_ot_s
    upsert_estudio(conn, REF, NOMBRE, FECHA, 'NACIONAL', 'Nacional',
                   r2(nat_mad   / nat_tot_s * 100),
                   r2(nat_cap_s / nat_tot_s * 100),
                   r2(nat_ot_s  / nat_tot_s * 100),
                   nat_ctrs)

    # ── Official results ──────────────────────────────────────────────────────
    # cols: cod_edo(0), maduro(12), capriles(13), sequera(14),
    #       bolivar(15), mora(16), mendez(17)
    wb_off = openpyxl.load_workbook(
        f'{DATA_2013_PRESIDENCIAL}/resultados oficiales elecciones presidenciales 2013.xlsx',
        read_only=True, data_only=True)
    ws_off = wb_off.active
    by_edo = defaultdict(lambda: {'mad': 0, 'cap': 0, 'otros': 0, 'tot': 0})
    nat_off = {'mad': 0, 'cap': 0, 'otros': 0, 'tot': 0}
    for row in ws_off.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        cod_edo = int(row[0])
        if cod_edo == 99:
            continue  # excluir exterior
        mad   = int(row[12] or 0)
        cap   = int(row[13] or 0)
        otros = (int(row[14] or 0) + int(row[15] or 0) +
                 int(row[16] or 0) + int(row[17] or 0))
        tot   = mad + cap + otros
        by_edo[cod_edo]['mad']   += mad
        by_edo[cod_edo]['cap']   += cap
        by_edo[cod_edo]['otros'] += otros
        by_edo[cod_edo]['tot']   += tot
        nat_off['mad']   += mad
        nat_off['cap']   += cap
        nat_off['otros'] += otros
        nat_off['tot']   += tot

    d = nat_off['tot']
    upsert_oficial(conn, REF, NOMBRE, FECHA, 'NACIONAL', 'Nacional',
                   r2(nat_off['mad'] / d * 100),
                   r2(nat_off['cap'] / d * 100),
                   r2(nat_off['otros'] / d * 100), d)
    for edo, s in by_edo.items():
        cod = f'{edo:02d}'
        if cod not in estado_nombre or s['tot'] == 0:
            continue
        upsert_oficial(conn, REF, NOMBRE, FECHA, cod, estado_nombre[cod],
                       r2(s['mad']   / s['tot'] * 100),
                       r2(s['cap']   / s['tot'] * 100),
                       r2(s['otros'] / s['tot'] * 100),
                       s['tot'])

    print(f'2013: estudio nacional Maduro={r2(nat_mad/nat_tot_s*100)}% Capriles={r2(nat_cap_s/nat_tot_s*100)}%'
          f' | oficial Maduro={r2(nat_off["mad"]/d*100)}% Capriles={r2(nat_off["cap"]/d*100)}%')


# ══════════════════════════════════════════════════════════════════════════════
def main():
    conn = get_conn()
    import_2012(conn)
    import_2013(conn)
    conn.commit()

    # ── Verify ────────────────────────────────────────────────────────────────
    print()
    for ref in ('2012-presidencial', '2013-presidencial'):
        ne = conn.execute("SELECT COUNT(*) FROM historico_estudios WHERE eleccion_ref=?", (ref,)).fetchone()[0]
        no = conn.execute("SELECT COUNT(*) FROM historico_oficial  WHERE eleccion_ref=?", (ref,)).fetchone()[0]
        nt = conn.execute("SELECT COUNT(*) FROM historico_estudios_turnos WHERE eleccion_ref=?", (ref,)).fetchone()[0]
        print(f'{ref}: estudios={ne} filas, oficial={no} filas, turnos={nt} filas')

    conn.close()
    print('\nImportación completada.')


if __name__ == '__main__':
    main()
