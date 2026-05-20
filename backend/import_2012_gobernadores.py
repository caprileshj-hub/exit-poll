"""
Import: Gobernadores 2012 - 23 exit polls regionales.

Lee data/2012/gobernadores/CORE4.xlsx y carga:
  - historico_estudios: resultados del exit poll por estado (ambito=slug)
  - historico_oficial: resultados oficiales por estado (Wikipedia/CNE)
  - historico_estudios_turnos: serie acumulativa por estado y turno

Uso: python import_2012_gobernadores.py [--reset]
"""
from __future__ import annotations

import json
import math
import os
import re
import sqlite3
import sys
import unicodedata
from collections import defaultdict

import openpyxl

DB = "exitpoll.db"
BASE = "data/2012/gobernadores"
FILE = "CORE4.xlsx"
REF = "2012-gobernadores"
FECHA = "2012-12-16"
NOMBRE = "Elecciones Regionales 2012"
ICC_REF = 0.04


SLUG_MAP = {
    "AM": "amazonas",
    "AN": "anzoategui",
    "AP": "apure",
    "AR": "aragua",
    "BA": "barinas",
    "BO": "bolivar",
    "CA": "carabobo",
    "CO": "cojedes",
    "DA": "delta-amacuro",
    "FA": "falcon",
    "GU": "guarico",
    "LA": "lara",
    "ME": "merida",
    "MI": "miranda",
    "MO": "monagas",
    "NE": "nueva-esparta",
    "PO": "portuguesa",
    "SU": "sucre",
    "TA": "tachira",
    "TR": "trujillo",
    "VA": "vargas",
    "YA": "yaracuy",
    "ZU": "zulia",
    "Amazonas": "amazonas",
    "Anzoategui": "anzoategui",
    "Apure": "apure",
    "Aragua": "aragua",
    "Barinas": "barinas",
    "Bolivar": "bolivar",
    "Carabobo": "carabobo",
    "Cojedes": "cojedes",
    "Delta Amacuro": "delta-amacuro",
    "Falcon": "falcon",
    "Guarico": "guarico",
    "Lara": "lara",
    "Merida": "merida",
    "Miranda": "miranda",
    "Monagas": "monagas",
    "Nueva Esparta": "nueva-esparta",
    "Portuguesa": "portuguesa",
    "Sucre": "sucre",
    "Tachira": "tachira",
    "Trujillo": "trujillo",
    "Vargas": "vargas",
    "Yaracuy": "yaracuy",
    "Zulia": "zulia",
}

NOMBRE_DISPLAY = {
    "amazonas": "Amazonas",
    "anzoategui": "Anzoategui",
    "apure": "Apure",
    "aragua": "Aragua",
    "barinas": "Barinas",
    "bolivar": "Bolivar",
    "carabobo": "Carabobo",
    "cojedes": "Cojedes",
    "delta-amacuro": "Delta Amacuro",
    "falcon": "Falcon",
    "guarico": "Guarico",
    "lara": "Lara",
    "merida": "Merida",
    "miranda": "Miranda",
    "monagas": "Monagas",
    "nueva-esparta": "Nueva Esparta",
    "portuguesa": "Portuguesa",
    "sucre": "Sucre",
    "tachira": "Tachira",
    "trujillo": "Trujillo",
    "vargas": "Vargas",
    "yaracuy": "Yaracuy",
    "zulia": "Zulia",
}

REGION_MAP = {
    "tachira": "Andina",
    "merida": "Andina",
    "trujillo": "Andina",
    "barinas": "Andina",
    "miranda": "Capital",
    "vargas": "Capital",
    "aragua": "Central",
    "cojedes": "Central",
    "carabobo": "Central",
    "falcon": "Centro Occidental",
    "lara": "Centro Occidental",
    "portuguesa": "Centro Occidental",
    "yaracuy": "Centro Occidental",
    "apure": "Guayana y Llanos",
    "bolivar": "Guayana y Llanos",
    "delta-amacuro": "Guayana y Llanos",
    "guarico": "Guayana y Llanos",
    "amazonas": "Guayana y Llanos",
    "nueva-esparta": "Nororiental e Insular",
    "monagas": "Nororiental e Insular",
    "sucre": "Nororiental e Insular",
    "anzoategui": "Nororiental e Insular",
    "zulia": "Zuliana",
}

TIPO_MAP = {}

OFICIALES = {
    "amazonas": {"o_gov": 37.93, "o_opos": 56.07, "o_otros": 5.10,
                 "ganador_gov": False, "cand_gov": "Nicia Maldonado (PSUV)",
                 "cand_opos": "Liborio Guarulla (MUD/MPV)"},
    "anzoategui": {"o_gov": 56.34, "o_opos": 42.98, "o_otros": 0.38,
                   "ganador_gov": True, "cand_gov": "Aristobulo Isturiz (PSUV)",
                   "cand_opos": "Antonio Barreto Sira (MUD/AD)"},
    "apure": {"o_gov": 63.02, "o_opos": 22.62, "o_otros": 14.12,
              "ganador_gov": True, "cand_gov": "Ramon Carrizalez (PSUV)",
              "cand_opos": "Luis Lippa (MUD)"},
    "aragua": {"o_gov": 55.54, "o_opos": 44.20, "o_otros": 0.15,
               "ganador_gov": True, "cand_gov": "Tareck El Aissami (PSUV)",
               "cand_opos": "Richard Mardo (MUD/PJ)"},
    "barinas": {"o_gov": 57.67, "o_opos": 42.29, "o_otros": 0.03,
                "ganador_gov": True, "cand_gov": "Adan Chavez (PSUV)",
                "cand_opos": "Julio Cesar Reyes (MUD/AP)"},
    "bolivar": {"o_gov": 46.03, "o_opos": 44.17, "o_otros": 8.17,
                "ganador_gov": True, "cand_gov": "Francisco Rangel Gomez (PSUV)",
                "cand_opos": "Andres Velasquez (MUD/LCR)",
                "nota_bolivar": "Resultado oficial impugnado publicamente por Andres Velasquez; se mantiene la categorizacion oficial CNE/Wikipedia."},
    "carabobo": {"o_gov": 55.48, "o_opos": 43.86, "o_otros": 0.21,
                 "ganador_gov": True, "cand_gov": "Francisco Ameliach (PSUV)",
                 "cand_opos": "Henrique Salas Feo (MUD/PV)"},
    "cojedes": {"o_gov": 63.38, "o_opos": 36.08, "o_otros": 0.18,
                "ganador_gov": True, "cand_gov": "Erika Farias (PSUV)",
                "cand_opos": "Alberto Galindez (MUD/AD)"},
    "delta-amacuro": {"o_gov": 77.24, "o_opos": 20.36, "o_otros": 2.31,
                      "ganador_gov": True, "cand_gov": "Lizeta Hernandez (PSUV)",
                      "cand_opos": "Arevalo Salazar (MUD/MAS)"},
    "falcon": {"o_gov": 51.50, "o_opos": 36.02, "o_otros": 11.85,
               "ganador_gov": True, "cand_gov": "Stella Lugo (PSUV)",
               "cand_opos": "Gregorio Graterol (MUD)"},
    "guarico": {"o_gov": 74.66, "o_opos": 25.28, "o_otros": 0.04,
                "ganador_gov": True, "cand_gov": "Ramon Rodriguez Chacin (PSUV)",
                "cand_opos": "Jose Manuel Gonzalez (MUD/AD)"},
    "lara": {"o_gov": 45.08, "o_opos": 54.66, "o_otros": 0.14,
             "ganador_gov": False, "cand_gov": "Luis Reyes Reyes (PSUV)",
             "cand_opos": "Henri Falcon (MUD/AP)"},
    "merida": {"o_gov": 50.18, "o_opos": 38.83, "o_otros": 10.61,
               "ganador_gov": True, "cand_gov": "Alexis Ramirez (PSUV)",
               "cand_opos": "Lester Rodriguez (MUD/COPEI)"},
    "miranda": {"o_gov": 47.71, "o_opos": 51.94, "o_otros": 0.31,
                "ganador_gov": False, "cand_gov": "Elias Jaua (PSUV)",
                "cand_opos": "Henrique Capriles (MUD/PJ)"},
    "monagas": {"o_gov": 54.93, "o_opos": 2.43, "o_otros": 41.86,
                "ganador_gov": True, "cand_gov": "Yelitze Santaella (PSUV)",
                "cand_opos": "Soraya Hernandez (MUD)",
                "nota_monagas": "Jose Gregorio Briceño obtuvo 41.86% como candidato fuera de GPP/MUD; se clasifica en otros."},
    "nueva-esparta": {"o_gov": 54.06, "o_opos": 45.72, "o_otros": 0.14,
                      "ganador_gov": True, "cand_gov": "Carlos Mata Figueroa (PSUV)",
                      "cand_opos": "Morel Rodriguez (MUD/AD)"},
    "portuguesa": {"o_gov": 53.96, "o_opos": 21.06, "o_otros": 24.51,
                   "ganador_gov": True, "cand_gov": "Wilmar Castro (PSUV)",
                   "cand_opos": "Ivan Colmenares (MUD)",
                   "nota_portuguesa": "Oswaldo Zerpa obtuvo 24.51% fuera de GPP/MUD; se clasifica en otros."},
    "sucre": {"o_gov": 59.29, "o_opos": 36.07, "o_otros": 3.46,
              "ganador_gov": True, "cand_gov": "Luis Acuna (PSUV)",
              "cand_opos": "Hernan Nunez (MUD)"},
    "tachira": {"o_gov": 53.96, "o_opos": 45.52, "o_otros": 0.36,
                "ganador_gov": True, "cand_gov": "Jose Vielma Mora (PSUV)",
                "cand_opos": "Cesar Perez Vivas (MUD/COPEI)"},
    "trujillo": {"o_gov": 82.22, "o_opos": 17.34, "o_otros": 0.42,
                 "ganador_gov": True, "cand_gov": "Henry Rangel Silva (PSUV)",
                 "cand_opos": "Jose Hernandez (MUD)"},
    "vargas": {"o_gov": 73.38, "o_opos": 25.22, "o_otros": 1.20,
               "ganador_gov": True, "cand_gov": "Jorge Garcia Carneiro (PSUV)",
               "cand_opos": "Jose Manuel Olivares (MUD/PJ)"},
    "yaracuy": {"o_gov": 61.46, "o_opos": 37.78, "o_otros": 0.31,
                "ganador_gov": True, "cand_gov": "Julio Leon Heredia (PSUV)",
                "cand_opos": "Biagio Pilieri (MUD/Convergencia)"},
    "zulia": {"o_gov": 52.19, "o_opos": 47.71, "o_otros": 0.04,
              "ganador_gov": True, "cand_gov": "Francisco Arias Cardenas (PSUV)",
              "cand_opos": "Pablo Perez Alvarez (MUD/UNT)"},
}


def _slug(name: str) -> str:
    return SLUG_MAP.get(name, re.sub(r"[^a-z0-9-]", "", unicodedata.normalize("NFD", name.lower()).encode("ascii", "ignore").decode()))


def _deff(n: int, k: int) -> float | None:
    if k <= 0 or n <= 0:
        return None
    m_bar = n / k
    return round(1 + (m_bar - 1) * ICC_REF, 2)


def _moe(n: int, deff: float | None) -> float | None:
    if not n:
        return None
    d = deff if deff else 1.0
    return round(1.96 * math.sqrt(d * 0.25 / n) * 100, 2)


def is_graphic_sheet(name: str, all_names: list[str]) -> bool:
    for n in all_names:
        if n == name:
            continue
        if name == "G" + n or name == "C" + n:
            return True
    return False


def parse_state_sheet(ws) -> dict | None:
    """
    CORE4.xlsx no trae el layout 2008 (Entidad/Centro/Turno/Cand...).
    Las hojas de estado estan agregadas por municipio: columnas B:E son
    Municipio/Candidato/Resultado/Recibidos y F:W son turnos 1..18. Para
    reutilizar las vistas existentes, cada municipio funciona como segmento
    de campo y se guarda una lectura acumulada por (municipio, turno).
    """
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    turnos: list[int] = []
    for i, row in enumerate(rows):
        if len(row) > 4 and row[1] == "Municipio" and row[2] == "Candidato":
            header_idx = i
            for v in row[5:]:
                if isinstance(v, (int, float)):
                    turnos.append(int(v))
            break
    if header_idx is None or not turnos:
        return None

    candidates = ["PSUV", "MUD", "Otro"]
    cand_idx = {c: i for i, c in enumerate(candidates)}
    by_mun: dict[int, dict[int, list[int]]] = defaultdict(lambda: defaultdict(lambda: [0] * len(candidates)))
    centros: set[int] = set()
    current_mun: int | None = None
    seq = 0

    for row in rows[header_idx + 1:]:
        cand = row[2] if len(row) > 2 else None
        if cand not in cand_idx:
            continue
        if cand == "PSUV":
            seq += 1
            raw_mun = row[1] if len(row) > 1 else None
            current_mun = int(raw_mun) if isinstance(raw_mun, (int, float)) else seq
            centros.add(current_mun)
        if current_mun is None:
            continue
        k = cand_idx[cand]
        for pos, turno in enumerate(turnos):
            col = 5 + pos
            val = row[col] if col < len(row) else None
            by_mun[current_mun][turno][k] += int(round(val)) if isinstance(val, (int, float)) else 0

    if not by_mun:
        return None

    readings: dict[tuple, list[int]] = {}
    total_votos = [0] * len(candidates)
    for centro, per_turno in by_mun.items():
        acum = [0] * len(candidates)
        for turno in sorted(turnos):
            vals = per_turno.get(turno, [0] * len(candidates))
            for k, val in enumerate(vals):
                acum[k] += val
            readings[(centro, turno)] = list(acum)
        for k, val in enumerate(acum):
            total_votos[k] += val

    total = sum(total_votos)
    pct = [round(v / total * 100, 2) if total else 0 for v in total_votos]

    return {
        "candidates": candidates,
        "centros": centros,
        "n_centros": len(centros),
        "max_turno": max(turnos),
        "total": total,
        "votos": total_votos,
        "pct": pct,
        "readings": readings,
    }


def build_turno_series(parsed: dict) -> list[dict]:
    """
    Para cada turno T, calcula el porcentaje acumulativo usando
    para cada centro su ultima lectura con turno <= T.
    """
    readings = parsed["readings"]
    max_turno = parsed["max_turno"]
    candidates = parsed["candidates"]
    n_cands = len(candidates)

    por_centro: dict[int, list[tuple]] = defaultdict(list)
    for (centro, turno), votos in readings.items():
        por_centro[centro].append((turno, votos))
    for c in por_centro:
        por_centro[c].sort()

    series = []
    for t in range(1, max_turno + 1):
        acum = [0] * n_cands
        centros_activos = 0
        for centro, lecturas in por_centro.items():
            ultima = None
            for turno, votos in lecturas:
                if turno <= t:
                    ultima = votos
                else:
                    break
            if ultima is not None:
                centros_activos += 1
                for k, v in enumerate(ultima):
                    acum[k] += v
        total_t = sum(acum)
        if total_t == 0:
            continue
        pct_t = [round(v / total_t * 100, 2) for v in acum]
        series.append({
            "turno": t,
            "pct_gov": pct_t[0] if pct_t else 0,
            "pct_opos": pct_t[1] if len(pct_t) > 1 else 0,
            "pct_otros": round(100 - pct_t[0] - (pct_t[1] if len(pct_t) > 1 else 0), 2),
            "num_centros": centros_activos,
        })
    return series


def main():
    reset = "--reset" in sys.argv
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = OFF")

    if reset:
        print("Eliminando datos previos de 2012-gobernadores...")
        conn.execute("DELETE FROM historico_estudios WHERE eleccion_ref=?", (REF,))
        conn.execute("DELETE FROM historico_oficial WHERE eleccion_ref=?", (REF,))
        conn.execute("DELETE FROM historico_estudios_turnos WHERE eleccion_ref=?", (REF,))
        conn.commit()

    estados: list[dict] = []
    path = os.path.join(BASE, FILE)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    state_sheets = [
        s for s in all_sheets
        if s in SLUG_MAP and not is_graphic_sheet(s, all_sheets)
    ]
    for sheet_name in state_sheets:
        slug = _slug(sheet_name)
        parsed = parse_state_sheet(wb[sheet_name])
        if not parsed:
            print(f"  [SKIP] {sheet_name}: sin datos")
            continue
        estados.append({"slug": slug, "sheet": sheet_name, "parsed": parsed})
    wb.close()

    print(f"Estados parseados: {len(estados)}")

    total_centros = sum(e["parsed"]["n_centros"] for e in estados)
    total_resp = sum(e["parsed"]["total"] for e in estados)
    gov_wins_e = sum(1 for e in estados if e["parsed"]["pct"] and len(e["parsed"]["pct"]) > 1
                     and e["parsed"]["pct"][0] > e["parsed"]["pct"][1])
    opos_wins_e = len(estados) - gov_wins_e

    notas_nac = json.dumps({
        "tipo": "coleccion_gobernadores",
        "n_estudios": len(estados),
        "n_centros_total": total_centros,
        "n_respondentes_total": total_resp,
        "icc_referencia": ICC_REF,
        "gov_wins_estudio": gov_wins_e,
        "opos_wins_estudio": opos_wins_e,
        "fuente_oficial": "Wikipedia/CNE 2012",
        "layout": "CORE4.xlsx hojas por estado agregadas por municipio/turno",
    }, ensure_ascii=False)

    conn.execute("""
        INSERT INTO historico_estudios
            (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
             pct_gov, pct_opos, pct_otros, num_centros, fuente, notas)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
            nombre=excluded.nombre, nombre_eleccion=excluded.nombre_eleccion,
            fecha_eleccion=excluded.fecha_eleccion, num_centros=excluded.num_centros,
            notas=excluded.notas, updated_at=datetime('now')
    """, (REF, "NACIONAL", "Gobernadores 2012 (coleccion)",
          NOMBRE, FECHA, 0, 0, 0, total_centros, "CORE4 Excel / importacion", notas_nac))

    for e in estados:
        slug = e["slug"]
        parsed = e["parsed"]
        cands = parsed["candidates"]
        pct = parsed["pct"]
        of_ = OFICIALES.get(slug, {})

        e_gov = pct[0] if pct else None
        e_opos = pct[1] if len(pct) > 1 else None
        e_otros = round(100 - (e_gov or 0) - (e_opos or 0), 2) if e_gov is not None else None

        deff = _deff(parsed["total"], parsed["n_centros"])
        moe_srs = _moe(parsed["total"], 1.0)
        moe_adj = _moe(parsed["total"], deff)

        notas_estado = json.dumps({
            "tipo": "gobernador",
            "tipo_cargo": TIPO_MAP.get(slug, "Gobernacion"),
            "region": REGION_MAP.get(slug, ""),
            "candidato_gov": cands[0] if cands else None,
            "candidato_opos": cands[1] if len(cands) > 1 else None,
            "candidato_otros": cands[2:] if len(cands) > 2 else [],
            "cand_gov_nombre": of_.get("cand_gov"),
            "cand_opos_nombre": of_.get("cand_opos"),
            "n_respondentes": parsed["total"],
            "deff_estimado": deff,
            "moe_srs_pp": moe_srs,
            "moe_ajustado_pp": moe_adj,
            "nota_bolivar": of_.get("nota_bolivar"),
            "nota_monagas": of_.get("nota_monagas"),
            "nota_portuguesa": of_.get("nota_portuguesa"),
            "layout": "municipios/segmentos agregados, no centros CNE crudos",
        }, ensure_ascii=False)

        conn.execute("""
            INSERT INTO historico_estudios
                (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
                 pct_gov, pct_opos, pct_otros, num_centros, fuente, notas)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
                nombre=excluded.nombre, nombre_eleccion=excluded.nombre_eleccion,
                fecha_eleccion=excluded.fecha_eleccion, pct_gov=excluded.pct_gov,
                pct_opos=excluded.pct_opos, pct_otros=excluded.pct_otros,
                num_centros=excluded.num_centros, notas=excluded.notas,
                updated_at=datetime('now')
        """, (REF, slug, NOMBRE_DISPLAY.get(slug, slug), NOMBRE, FECHA,
              e_gov, e_opos, e_otros, parsed["n_centros"],
              "CORE4 Excel / importacion", notas_estado))

        conn.execute("""
            INSERT INTO historico_oficial
                (eleccion_ref, ambito, nombre, nombre_eleccion, fecha_eleccion,
                 pct_gov, pct_opos, pct_otros, fuente)
            VALUES (?,?,?,?,?,?,?,?,?)
            ON CONFLICT(eleccion_ref, ambito) DO UPDATE SET
                nombre=excluded.nombre, nombre_eleccion=excluded.nombre_eleccion,
                fecha_eleccion=excluded.fecha_eleccion, pct_gov=excluded.pct_gov,
                pct_opos=excluded.pct_opos, pct_otros=excluded.pct_otros,
                updated_at=datetime('now')
        """, (REF, slug, NOMBRE_DISPLAY.get(slug, slug), NOMBRE, FECHA,
              of_.get("o_gov", 0), of_.get("o_opos", 0), of_.get("o_otros", 0),
              "Wikipedia/CNE 2012"))

        serie = build_turno_series(parsed)
        for s in serie:
            conn.execute("""
                INSERT INTO historico_estudios_turnos
                    (eleccion_ref, ambito, turno, pct_gov, pct_opos, pct_otros, num_centros)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(eleccion_ref, ambito, turno) DO UPDATE SET
                    pct_gov=excluded.pct_gov, pct_opos=excluded.pct_opos,
                    pct_otros=excluded.pct_otros, num_centros=excluded.num_centros,
                    updated_at=datetime('now')
            """, (REF, slug, s["turno"], s["pct_gov"], s["pct_opos"], s["pct_otros"], s["num_centros"]))

        gov_str = f"{e_gov}%" if e_gov is not None else "-"
        opos_str = f"{e_opos}%" if e_opos is not None else "-"
        win_e = "GOV " if (e_gov and e_opos and e_gov > e_opos) else "OPOS"
        win_o = "GOV " if of_.get("ganador_gov") else ("OPOS" if of_.get("ganador_gov") is False else "  ? ")
        ok = "OK" if (win_e.strip() == win_o.strip()) else "XX"
        print(f"  {ok} {slug:<18} estudio={gov_str:>6}/{opos_str:<6} {win_e}  "
              f"oficial={of_.get('o_gov','?')}%/{of_.get('o_opos','?')}% {win_o}  "
              f"n={parsed['total']:>6}  k={parsed['n_centros']:>3}  "
              f"T={parsed['max_turno']:>2}  DEFF={deff}")

    conn.commit()
    conn.execute("PRAGMA foreign_keys = ON")
    conn.close()

    print()
    ne = len(estados) + 1
    gov_wins_o = sum(1 for e in estados if OFICIALES.get(e["slug"], {}).get("ganador_gov") is True)
    correctos = sum(1 for e in estados
                    if e["parsed"]["pct"] and len(e["parsed"]["pct"]) > 1
                    and OFICIALES.get(e["slug"], {}).get("ganador_gov") is not None
                    and (e["parsed"]["pct"][0] > e["parsed"]["pct"][1]) == OFICIALES[e["slug"]]["ganador_gov"])
    con_oficial = sum(1 for e in estados if OFICIALES.get(e["slug"], {}).get("ganador_gov") is not None)
    print(f"Filas cargadas: {ne} estudio + {len(estados)} oficial + turnos")
    print(f"Estudio: gov gana en {gov_wins_e}/{len(estados)}")
    print(f"Oficial: gov gana en {gov_wins_o}/{con_oficial}")
    print(f"Ganador correcto: {correctos}/{con_oficial}  ({round(correctos/con_oficial*100, 1)}%)")


if __name__ == "__main__":
    main()
